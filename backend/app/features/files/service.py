"""Upload a file → one of two single-purpose imports (see save_and_import `mode`):

- "project_db": import the table into the project DB — a real project table (queried via the
  primary adapter). No disk file, no `files` row.
- "excel":      keep the original workbook on the shared volume so the excel-server (HTTP) can
  read/edit it. No SQL import. Recorded as a `files` row with only the disk path.
"""
from __future__ import annotations

import asyncio
import csv
import logging
import re
import shutil
import uuid
from io import BytesIO
from pathlib import Path

from app.config import get_settings
from app.features.files import repository as repo

logger = logging.getLogger("files")

# Delimited text (CSV/TSV/delimited TXT) + the Excel-family formats calamine/openpyxl/xlrd read.
_DELIMITED_EXTS = (".csv", ".tsv", ".txt")
_EXCEL_EXTS = (".xlsx", ".xlsm", ".xls", ".xlsb", ".ods", ".xltx", ".xltm")
# Of the Excel family, only zip-XML workbooks stream via openpyxl read-only (DB imports);
# the rest must be fully loaded to parse (router caps those at 30 MB).
_STREAM_XLSX_EXTS = (".xlsx", ".xlsm")
_TABULAR_EXTS = _DELIMITED_EXTS + _EXCEL_EXTS
# Anything tabular can be served to the Excel tools (converted to .xlsx if not already).
_EXCEL_EDITABLE = _TABULAR_EXTS


class FileImportError(Exception):
    """Raised when an upload cannot be imported (e.g. non-tabular file into a DB)."""


def _project_table_name(filename: str, sheet: str | None = None) -> str:
    """Clean, file-derived table name for a project-DB import (NO t_ prefix — that prefix is
    reserved for session/file tables so the request router can tell the two apart). Ensures the
    name is a valid identifier (won't start with a digit)."""
    stem = Path(filename).stem.lower()
    slug = re.sub(r"[^a-z0-9_]+", "_", stem).strip("_") or "data"
    name = slug
    if sheet is not None:
        ssl = re.sub(r"[^a-z0-9_]+", "_", str(sheet).lower()).strip("_") or "sheet"
        name = f"{name}_{ssl}"
    if name[0].isdigit():
        name = f"_{name}"
    return name[:63]


def _sanitize_columns(df):
    """Unique, SQLite-safe column names (duplicate Excel headers are common)."""
    seen: dict[str, int] = {}
    cols: list[str] = []
    for raw in df.columns:
        if isinstance(raw, tuple):
            base = "_".join(str(p).strip() for p in raw if str(p).strip()).strip("_")
        else:
            base = str(raw).strip() if raw is not None else ""
        base = re.sub(r"\s+", " ", base) or "column"
        if base.lower() == "nan":
            base = "column"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        cols.append(base[:120])
    df = df.copy()
    df.columns = cols
    return df


def _safe_sheet_name(name: object) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", str(name)).strip()[:31] or "Sheet1"


def _safe_filename(filename: str, *, suffix: str = "") -> str:
    """Make a client-supplied filename safe to use as a single path segment.

    Strips any directory component (so '../../x' → 'x') and unsafe characters, so it can never
    escape the target directory. The caller still verifies containment as a second guard."""
    base = re.split(r"[\\/]", filename)[-1]             # last segment (handles / and \)
    base = re.sub(r"[^A-Za-z0-9._ \-()]+", "_", base)
    base = base.replace("..", "").strip(". ") or "upload"   # no parent refs, no leading/trailing dots
    if suffix and not base.lower().endswith(suffix):
        base = f"{Path(base).stem or 'upload'}{suffix}"
    return base[:200]


def _looks_european(sample: str) -> bool:
    return bool(re.search(r"\d+,\d{1,2}(?:[;\s]|$)", sample))


_SNIFF_BYTES = 65536  # sniff encoding/delimiter/decimal from the first 64 KB only


def _sniff_csv(path: Path, ext: str) -> dict:
    """Detect encoding + delimiter + decimal style from the head of the file and
    return kwargs for pd.read_csv. Sniffing a 64 KB sample keeps RAM flat no
    matter how large the file is."""
    with path.open("rb") as fh:
        head = fh.read(_SNIFF_BYTES)

    if head.startswith(b"\xef\xbb\xbf"):
        encoding = "utf-8-sig"
    else:
        encoding = "latin-1"
        for enc in ("utf-8", "cp1258", "latin-1"):
            try:
                head.decode(enc)
                encoding = enc
                break
            except UnicodeDecodeError:
                continue

    sample = head[:8192].decode(encoding, errors="replace")
    if ext == ".txt" and not any(d in sample for d in (",", ";", "\t")):
        raise FileImportError("Text file is not delimited — cannot import as a table")
    try:
        sep = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        sep = "\t" if ext == ".tsv" else ","

    decimal = "," if sep == ";" and _looks_european(sample) else "."
    thousands = "." if decimal == "," else ","
    return {"sep": sep, "encoding": encoding, "decimal": decimal, "thousands": thousands}


def _read_csv_full(path: Path, ext: str):
    """Whole-file CSV read (small files / xlsx conversion). Streaming imports use
    _sniff_csv + pd.read_csv(chunksize=...) directly — see _import_csv_streaming."""
    import pandas as pd  # lazy

    return pd.read_csv(path, **_sniff_csv(path, ext))


def _read_excel_sheets(path: Path) -> dict:
    """All sheets as {name: df}. calamine first (xlsx/xls/xlsb/ods), then pandas default.
    NOTE: loads the whole workbook — callers must enforce the Excel size cap first."""
    import pandas as pd  # lazy

    last_err: Exception | None = None
    for engine in ("calamine", None):
        try:
            kwargs: dict = {"sheet_name": None}
            if engine:
                kwargs["engine"] = engine
            return pd.read_excel(path, **kwargs)
        except Exception as e:  # noqa: BLE001
            last_err = e
            logger.warning("excel read (engine=%s) failed: %s", engine, e)
    raise FileImportError(f"Could not read spreadsheet ({last_err})")


def _read_tables(ext: str, path: Path) -> dict:
    """Parse a tabular FILE into {sheet_label_or_None: DataFrame}. Delimited / single-sheet
    workbooks → key None (table name carries no sheet suffix); a multi-sheet workbook →
    one entry per sheet (each becomes its own table). Reads from disk, not a bytes blob."""
    if ext in _DELIMITED_EXTS:
        return {None: _sanitize_columns(_read_csv_full(path, ext))}
    if ext in _EXCEL_EXTS:
        sheets = _read_excel_sheets(path)
        usable = {
            str(name): _sanitize_columns(df)
            for name, df in sheets.items()
            if str(name).strip() and getattr(df, "shape", (0, 0))[1] > 0
        }
        if not usable:
            raise FileImportError("Spreadsheet has no readable sheets")
        if len(usable) == 1:
            return {None: next(iter(usable.values()))}
        return usable
    raise FileImportError(f"Unsupported file type: {ext}")


def _to_xlsx_bytes(ext: str, path: Path) -> bytes:
    """Convert a tabular file to .xlsx bytes so the (xlsx-only) excel-server can read it.
    Delimited → one sheet; Excel-family → every sheet preserved (formulas are not).

    Written via openpyxl directly rather than pandas' ExcelWriter, which is broken on
    pandas 3.0 + openpyxl 3.1 ("At least one sheet must be visible")."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if ext in _DELIMITED_EXTS:
        sheets = {"Sheet1": _read_csv_full(path, ext)}
    else:
        sheets = _read_excel_sheets(path)

    wb = Workbook()
    default_ws = wb.active  # may be None per the type stubs
    if default_ws is not None:
        wb.remove(default_ws)  # drop the default empty sheet
    for name, df in sheets.items():
        ws = wb.create_sheet(_safe_sheet_name(name))
        for row in dataframe_to_rows(_sanitize_columns(df), index=False, header=True):
            ws.append(row)
    if not wb.sheetnames:  # nothing parsed → keep a valid (visible) workbook
        wb.create_sheet("Sheet1")
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _session_uploads_dir(session_id: str) -> Path:
    """Where this session's Excel workbooks live (shared with the excel-server)."""
    return Path(get_settings().data_root) / "uploads" / session_id


async def excel_file_paths(session_id: str, file_ids: list[str] | None = None) -> list[str]:
    """Relative paths ('<session>/<file>') of workbooks the excel-server can open for this session —
    injected into the Excel agent prompt so it opens/saves the RIGHT file under the session folder.

    STRICT scoping: if `file_ids` is given (the ids the user picked this turn), return ONLY the
    workbooks among them — an empty result when no workbook is picked (the caller treats this as
    "not an Excel turn"). `file_ids=None` returns ALL of the session's workbooks."""
    rows = await repo.list_for_session(session_id)
    workbooks = [r for r in rows if r.get("disk_path")]
    if file_ids is not None:
        wanted = set(file_ids)
        workbooks = [r for r in workbooks if str(r.get("id")) in wanted]
    return [f"{session_id}/{Path(r['disk_path']).name}" for r in workbooks]


def snapshot_excel_dir(session_id: str) -> dict[str, float]:
    """{filename: mtime} of the session's .xlsx files BEFORE an excel turn, used to detect which
    workbook(s) the agent created or modified."""
    d = _session_uploads_dir(session_id)
    if not d.exists():
        return {}
    return {p.name: p.stat().st_mtime for p in d.glob("*.xlsx") if p.is_file()}


async def register_excel_outputs(
    user_id: str, session_id: str, before: dict[str, float]
) -> list[dict]:
    """After an excel turn: for each .xlsx the agent created or modified, ensure a `files` row
    exists and return a `file_export` tool_event so the FE shows an inline download card."""
    d = _session_uploads_dir(session_id)
    if not d.exists():
        return []
    by_path = {r.get("disk_path"): r for r in await repo.list_for_session(session_id)}
    out: list[dict] = []
    for p in sorted(d.glob("*.xlsx")):
        if not p.is_file():
            continue
        prev = before.get(p.name)
        if prev is not None and p.stat().st_mtime <= prev + 1e-6:
            continue  # untouched this turn
        row = by_path.get(str(p))
        if row is None:  # a NEW workbook the agent created → register it so it is listable/downloadable
            row = await repo.insert_file(user_id, session_id, p.name, str(p), p.stat().st_size)
        out.append({
            "tool": "excel", "type": "file_export",
            "payload": {"filename": p.name, "sessionFileId": row["id"]},
        })
    return out


async def save_and_import(
    user_id: str,
    session_id: str,
    filename: str,
    path: Path,
    *,
    mode: str,
    project_id: str | None = None,
    project_db_url: str | None = None,
    target_table: str | None = None,
) -> dict:
    """Single-purpose import per `mode`. `path` is the spooled upload on disk (the router
    owns and deletes it). Any failure raises FileImportError → 400:
    - "project_db": import into the project DB. By default a NEW table (file-named); when
      `target_table` is given, APPEND the rows into that existing table instead.
    - "excel":      keep the original on the shared volume for the excel-server (no SQL import).
    """
    logger.info("→ save_and_import(user_id=%r session_id=%r filename=%r mode=%r project_id=%r target_table=%r size=%d)", user_id, session_id, filename, mode, project_id, target_table, path.stat().st_size)  # autolog
    if mode == "project_db":
        if target_table:
            return await _append_to_project_table(filename, path, project_id, project_db_url, target_table)
        return await _import_to_project_db(filename, path, project_id, project_db_url)
    if mode == "excel":
        return await _save_for_excel(user_id, session_id, filename, path)
    raise FileImportError(f"Unknown import mode: {mode!r}")


async def _import_to_project_db(
    filename: str, path: Path, project_id: str | None, project_db_url: str | None
) -> dict:
    """Import into the project DB → a real project table (queried via the pooled primary
    adapter, so chat sees it immediately). Nothing on disk, no `files` row — synthesized meta."""
    if not (project_id and project_db_url):
        raise FileImportError("No project database to import into")
    ext = Path(filename).suffix.lower()
    if ext not in _TABULAR_EXTS:
        raise FileImportError("Only CSV/Excel files can be imported into the database")
    from app.agent.pool import get_connection_pool

    adapter = await get_connection_pool().adapter_for(project_id, project_db_url)

    if ext in _DELIMITED_EXTS:  # CSV/TSV/TXT stream — flat RAM at any size
        return await _import_csv_streaming(filename, path, ext, adapter)
    if ext in _STREAM_XLSX_EXTS:  # xlsx/xlsm stream via openpyxl read-only — flat RAM
        return await _import_xlsx_streaming(filename, path, adapter)

    # Legacy Excel-family: whole-workbook load (router caps these at 30 MB), off the event loop.
    tables = await asyncio.to_thread(_read_tables, ext, path)
    # Project tables keep a clean, file-derived name (no t_). Refuse to clobber an existing table —
    # the user must rename the file or drop the table first (no silent overwrite of real data).
    targets = {label: _project_table_name(filename, label) for label in tables}
    existing = set((await adapter.get_schema()).keys())
    clash = sorted({t for t in targets.values() if t in existing})
    if clash:
        raise FileImportError(
            f"Table already exists in the database: {', '.join(clash)}. "
            "Rename the file or drop the existing table first."
        )
    created: list[dict] = []
    for label, df in tables.items():
        try:
            await adapter.import_dataframe(targets[label], df, if_exists="fail")
        except ValueError as e:  # raced with a concurrent import that just created the table
            raise FileImportError(f"Table '{targets[label]}' already exists.") from e
        created.append({"name": targets[label], "columns": [str(c) for c in df.columns]})
    return {
        "id": str(uuid.uuid4()),  # synthesized: nothing persisted in `files`
        "filename": filename,
        "size_bytes": path.stat().st_size,
        "created_at": None,
        # New project tables → the FE offers a "describe this table" step (data dictionary).
        "tables": created,
    }


_CSV_CHUNK_ROWS = 50_000  # ~50k rows per DB round-trip keeps peak RAM at tens of MB


async def _import_csv_streaming(filename: str, path: Path, ext: str, adapter) -> dict:
    """Stream a delimited file into a NEW project table without ever loading the
    whole file: pandas reads `_CSV_CHUNK_ROWS` at a time (in a worker thread so the
    event loop stays free) and each chunk is appended to the table. On a mid-stream
    failure the half-written table is dropped — no partial imports survive."""
    import pandas as pd  # lazy

    table = _project_table_name(filename)
    if table in set((await adapter.get_schema()).keys()):
        raise FileImportError(
            f"Table already exists in the database: {table}. "
            "Rename the file or drop the existing table first."
        )

    kwargs = _sniff_csv(path, ext)
    reader = pd.read_csv(path, chunksize=_CSV_CHUNK_ROWS, **kwargs)

    def _next_chunk():
        try:
            return next(reader)
        except StopIteration:
            return None

    columns: list[str] = []
    total_rows = 0
    if_exists = "fail"  # first chunk creates the table; ValueError = lost a creation race
    try:
        while (chunk := await asyncio.to_thread(_next_chunk)) is not None:
            df = _sanitize_columns(chunk)  # same headers every chunk → same names
            if not columns:
                columns = [str(c) for c in df.columns]
            try:
                await adapter.import_dataframe(table, df, if_exists=if_exists)
            except ValueError as e:
                raise FileImportError(f"Table '{table}' already exists.") from e
            if_exists = "append"
            total_rows += len(df)
    except FileImportError:
        raise
    except Exception as e:  # noqa: BLE001 — clean up the partial table, surface a clean reason
        if total_rows or if_exists == "append":
            try:
                await adapter.execute(f'DROP TABLE IF EXISTS "{table}"')
            except Exception:  # noqa: BLE001
                logger.warning("could not drop partial table %s", table)
        raise FileImportError(f"Import failed after {total_rows} rows: {e}") from e
    finally:
        close = getattr(reader, "close", None)
        if close:
            close()

    if total_rows == 0 and not columns:
        raise FileImportError("File contains no rows")
    return {
        "id": str(uuid.uuid4()),  # synthesized: nothing persisted in `files`
        "filename": filename,
        "size_bytes": path.stat().st_size,
        "created_at": None,
        # New project tables → the FE offers a "describe this table" step (data dictionary).
        "tables": [{"name": table, "columns": columns}],
    }


_SHARED_STRINGS_MAX = 64 * 1024 * 1024  # 64 MB uncompressed — the one xlsx part that can't stream


def _check_xlsx_streamable(path: Path) -> None:
    """xlsx streams row-by-row EXCEPT shared strings (the workbook's unique text), which
    openpyxl must hold fully in RAM. The zip directory declares that part's uncompressed
    size — reject text-heavy monsters in a few ms, before any parsing."""
    import zipfile

    try:
        with zipfile.ZipFile(path) as z:
            try:
                info = z.getinfo("xl/sharedStrings.xml")
            except KeyError:
                return  # numeric-only workbook — nothing big to hold in RAM
            if info.file_size > _SHARED_STRINGS_MAX:
                raise FileImportError(
                    "This workbook contains too much unique text to import safely. "
                    "Export the data as CSV and upload that instead."
                )
    except zipfile.BadZipFile as e:
        raise FileImportError("Not a valid Excel (.xlsx) file") from e


def _xlsx_header_row(ws) -> tuple[int, list] | None:
    """(row_index, raw_values) of the first non-empty row of a read-only worksheet —
    the header (mirrors pandas, which skips leading blank rows). None = empty sheet.

    MUST be called after ws.reset_dimensions(): some writers declare a bogus dimension
    (up to A1:XFD1048576) and openpyxl would otherwise yield millions of empty rows of
    16384 Nones each — which looks like a hang. reset_dimensions() makes iteration
    follow the sheet's ACTUAL content instead of the declared bounds."""
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None and str(v).strip() != "" for v in row):
            raw = list(row)
            while raw and raw[-1] is None:  # styled-but-empty trailing cells
                raw.pop()
            return (idx, raw) if raw else None
    return None


def _reset_ws_dimensions(wb) -> None:
    """Ignore each sheet's declared dimension (writers lie — see _xlsx_header_row)."""
    for ws in wb.worksheets:
        reset = getattr(ws, "reset_dimensions", None)
        if reset:
            reset()


def _xlsx_columns(raw_header: list) -> list[str]:
    """Sanitized, de-duplicated column names from a raw header row (same rules as the
    whole-load path, so streamed and non-streamed imports name columns identically)."""
    import pandas as pd  # lazy

    return [str(c) for c in _sanitize_columns(pd.DataFrame(columns=raw_header)).columns]


def _pull_xlsx_chunk(rows_iter, width: int, limit: int) -> list[list]:
    """Up to `limit` data rows from a read-only sheet iterator, normalized to `width`
    cells (extra cells dropped, short rows padded). Fully-empty rows are skipped.
    Runs in a worker thread — this is the blocking XML-parsing part."""
    out: list[list] = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        r = list(row[:width])
        if len(r) < width:
            r += [None] * (width - len(r))
        out.append(r)
        if len(out) >= limit:
            break
    return out


async def _import_xlsx_streaming(filename: str, path: Path, adapter) -> dict:
    """Stream an .xlsx/.xlsm workbook into NEW project table(s): openpyxl's read-only
    mode SAX-parses rows forward, so peak RAM is one 50k-row chunk + shared strings
    (capped by _check_xlsx_streamable) — never the whole workbook. Mirrors
    _import_csv_streaming: on a mid-stream failure every table THIS import created is
    dropped (a table lost to a concurrent-creation race is not ours to drop)."""
    from openpyxl import load_workbook

    import pandas as pd  # lazy

    _check_xlsx_streamable(path)
    try:
        wb = await asyncio.to_thread(load_workbook, path, read_only=True, data_only=True)
    except FileImportError:
        raise
    except Exception as e:  # noqa: BLE001
        raise FileImportError(f"Could not read spreadsheet ({e})") from e

    try:
        _reset_ws_dimensions(wb)
        # Pre-pass: header row of every sheet (cheap — parses only the top of each sheet).
        headers: dict[str, tuple[int, list]] = {}
        for ws in wb.worksheets:
            found = await asyncio.to_thread(_xlsx_header_row, ws)
            if found is not None:
                headers[str(ws.title)] = found
        if not headers:
            raise FileImportError("Spreadsheet has no readable sheets")
        logger.info("xlsx streaming %r: %d usable sheet(s): %s", filename, len(headers), list(headers))

        # Same naming rules as the whole-load path: single sheet → file-derived name,
        # multi-sheet → per-sheet suffix. Refuse to clobber existing tables (checked
        # for ALL sheets up front, so a multi-sheet import is all-or-nothing).
        single = len(headers) == 1
        targets = {name: _project_table_name(filename, None if single else name) for name in headers}
        existing = set((await adapter.get_schema()).keys())
        clash = sorted({t for t in targets.values() if t in existing})
        if clash:
            raise FileImportError(
                f"Table already exists in the database: {', '.join(clash)}. "
                "Rename the file or drop the existing table first."
            )

        created: list[dict] = []
        cur_created: str | None = None
        try:
            for name, (hdr_idx, raw_header) in headers.items():
                table = targets[name]
                cols = _xlsx_columns(raw_header)
                rows_iter = wb[name].iter_rows(min_row=hdr_idx + 1, values_only=True)
                if_exists = "fail"  # first chunk creates the table; ValueError = creation race
                cur_created = None
                total = 0
                while True:
                    rows = await asyncio.to_thread(_pull_xlsx_chunk, rows_iter, len(cols), _CSV_CHUNK_ROWS)
                    df = pd.DataFrame(rows, columns=cols)
                    try:
                        await adapter.import_dataframe(table, df, if_exists=if_exists)
                    except ValueError as e:  # raced — the existing table is NOT ours to drop
                        raise FileImportError(f"Table '{table}' already exists.") from e
                    if if_exists == "fail":
                        cur_created = table
                    if_exists = "append"
                    total += len(rows)
                    logger.info("xlsx streaming %r → %s: %d rows so far", filename, table, total)
                    if len(rows) < _CSV_CHUNK_ROWS:
                        break
                created.append({"name": table, "columns": cols})
                cur_created = None
        except BaseException as e:
            partial = [c["name"] for c in created] + ([cur_created] if cur_created else [])
            for t in partial:
                try:
                    await adapter.execute(f'DROP TABLE IF EXISTS "{t}"')
                except Exception:  # noqa: BLE001
                    logger.warning("could not drop partial table %s", t)
            if isinstance(e, FileImportError):
                raise
            raise FileImportError(f"Import failed: {e}") from e
    finally:
        wb.close()

    return {
        "id": str(uuid.uuid4()),  # synthesized: nothing persisted in `files`
        "filename": filename,
        "size_bytes": path.stat().st_size,
        "created_at": None,
        # New project tables → the FE offers a "describe this table" step (data dictionary).
        "tables": created,
    }


def _align_to_table(df, table_columns: list) -> tuple:
    """Match the uploaded DataFrame's columns to an existing table's columns by name (flexible
    import). Returns (aligned_df, missing_required) where:
      - aligned_df keeps only columns present in BOTH (renamed to the table's exact casing),
        so extra file columns are dropped and the DB fills any column the file omits;
      - missing_required lists table columns the file lacks that are NOT NULL with no default
        (and not the auto-increment PK) — i.e. ones the append would fail on.
    Matching is case-insensitive on the column name (the file's headers were already sanitized)."""
    by_lower = {str(c.name).lower(): c for c in table_columns}
    rename: dict = {}
    for raw in df.columns:
        hit = by_lower.get(str(raw).lower())
        if hit is not None:
            rename[raw] = hit.name
    aligned = df[list(rename.keys())].rename(columns=rename) if rename else df.iloc[:, :0]
    present = {c.name for c in table_columns if c.name in set(rename.values())}
    missing_required = [
        c.name for c in table_columns
        if c.name not in present and not c.nullable and c.default is None and not c.pk
    ]
    return aligned, missing_required


async def _append_to_project_table(
    filename: str, path: Path, project_id: str | None, project_db_url: str | None,
    target_table: str,
) -> dict:
    """Append a file's rows INTO an existing project table (flexible column matching). Columns
    are matched to the target by name; extras are ignored and omitted columns fall back to the
    DB's default/NULL/auto-increment. CSV streams in chunks; Excel loads whole (router-capped).
    Raises FileImportError on any problem (→ 400).

    NOTE: appends are NOT rolled back on a mid-stream failure — the target table pre-existed
    with real data, so dropping it would be worse than leaving partially-appended rows."""
    if not (project_id and project_db_url):
        raise FileImportError("No project database to import into")
    ext = Path(filename).suffix.lower()
    if ext not in _TABULAR_EXTS:
        raise FileImportError("Only CSV/Excel files can be imported into the database")

    from app.agent.pool import get_connection_pool
    adapter = await get_connection_pool().adapter_for(project_id, project_db_url)
    schema = await adapter.get_schema()
    if target_table not in schema:
        raise FileImportError(f"Table '{target_table}' doesn't exist in the database.")

    def _validate(aligned, missing_required):
        if aligned.shape[1] == 0:
            raise FileImportError(
                f"None of the file's columns match the columns of '{target_table}'. "
                "Check the headers or import as a new table."
            )
        if missing_required:
            raise FileImportError(
                f"The file is missing required column(s) of '{target_table}': "
                f"{', '.join(missing_required)}."
            )

    if ext in _DELIMITED_EXTS:  # CSV/TSV/TXT — stream in chunks, flat RAM
        import pandas as pd  # lazy

        reader = pd.read_csv(path, chunksize=_CSV_CHUNK_ROWS, **_sniff_csv(path, ext))

        def _next_chunk():
            try:
                return next(reader)
            except StopIteration:
                return None

        checked = False
        try:
            while (chunk := await asyncio.to_thread(_next_chunk)) is not None:
                aligned, missing_required = _align_to_table(
                    _sanitize_columns(chunk), schema[target_table]
                )
                if not checked:  # headers identical every chunk → validate once
                    _validate(aligned, missing_required)
                    checked = True
                await adapter.import_dataframe(target_table, aligned, if_exists="append")
        except FileImportError:
            raise
        except Exception as e:  # noqa: BLE001 — surface a clean reason, not the raw driver text
            from app.agent.graph import dbtools
            raise FileImportError(
                f"Couldn’t append to '{target_table}': {dbtools.clean_db_error(str(e))}."
            ) from e
        finally:
            close = getattr(reader, "close", None)
            if close:
                close()
    elif ext in _STREAM_XLSX_EXTS:  # xlsx/xlsm — stream in chunks, flat RAM
        import pandas as pd  # lazy

        from openpyxl import load_workbook

        _check_xlsx_streamable(path)
        try:
            wb = await asyncio.to_thread(load_workbook, path, read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            raise FileImportError(f"Could not read spreadsheet ({e})") from e
        try:
            _reset_ws_dimensions(wb)
            usable: list[tuple[str, tuple[int, list]]] = []
            for ws in wb.worksheets:
                found = await asyncio.to_thread(_xlsx_header_row, ws)
                if found is not None:
                    usable.append((str(ws.title), found))
            if not usable:
                raise FileImportError("Spreadsheet has no readable sheets")
            if len(usable) > 1:
                raise FileImportError(
                    "This file has multiple sheets — appending supports a single table. "
                    "Split the sheets or import as a new table instead."
                )
            name, (hdr_idx, raw_header) = usable[0]
            cols = _xlsx_columns(raw_header)
            rows_iter = wb[name].iter_rows(min_row=hdr_idx + 1, values_only=True)
            checked = False
            total = 0
            while True:
                rows = await asyncio.to_thread(_pull_xlsx_chunk, rows_iter, len(cols), _CSV_CHUNK_ROWS)
                df = pd.DataFrame(rows, columns=cols)
                aligned, missing_required = _align_to_table(df, schema[target_table])
                if not checked:  # same header every chunk → validate once
                    _validate(aligned, missing_required)
                    checked = True
                if len(aligned):
                    try:
                        await adapter.import_dataframe(target_table, aligned, if_exists="append")
                    except Exception as e:  # noqa: BLE001 — surface a clean reason
                        from app.agent.graph import dbtools
                        raise FileImportError(
                            f"Couldn’t append to '{target_table}': {dbtools.clean_db_error(str(e))}."
                        ) from e
                total += len(rows)
                logger.info("xlsx append %r → %s: %d rows so far", filename, target_table, total)
                if len(rows) < _CSV_CHUNK_ROWS:
                    break
        finally:
            wb.close()
    else:  # legacy Excel-family: whole-workbook load (router caps at 30 MB), off the event loop
        tables = await asyncio.to_thread(_read_tables, ext, path)
        if len(tables) > 1:
            raise FileImportError(
                "This file has multiple sheets — appending supports a single table. "
                "Split the sheets or import as a new table instead."
            )
        df = next(iter(tables.values()))
        aligned, missing_required = _align_to_table(df, schema[target_table])
        _validate(aligned, missing_required)
        try:
            await adapter.import_dataframe(target_table, aligned, if_exists="append")
        except Exception as e:  # noqa: BLE001 — surface a clean reason, not the raw driver text
            from app.agent.graph import dbtools
            raise FileImportError(
                f"Couldn’t append to '{target_table}': {dbtools.clean_db_error(str(e))}."
            ) from e
    return {
        "id": str(uuid.uuid4()),  # synthesized: nothing persisted in `files`
        "filename": filename,
        "size_bytes": path.stat().st_size,
        "created_at": None,
    }


async def _save_for_excel(user_id: str, session_id: str, filename: str, path: Path) -> dict:
    """Keep a workbook on the shared volume so the excel-server (xlsx-only) can read/edit it.
    Native .xlsx is COPIED as-is (never parsed — preserves sheets/formulas, zero RAM); other
    tabular formats are converted to .xlsx (whole-load, router-capped at 30 MB). No SQL
    import — recorded as a `files` row with only the disk path."""
    ext = Path(filename).suffix.lower()
    if ext not in _EXCEL_EDITABLE:
        raise FileImportError("This file type can't be opened as an Excel workbook")
    out_name = _safe_filename(filename, suffix=".xlsx")
    uploads = (Path(get_settings().data_root) / "uploads" / session_id).resolve()
    uploads.mkdir(parents=True, exist_ok=True)
    disk_path = (uploads / out_name).resolve()
    # Second guard: the resolved path MUST stay inside uploads/<session>/ (defence in depth
    # against a crafted filename that survived sanitisation).
    if uploads not in disk_path.parents:
        raise FileImportError("Invalid file name")
    if ext == ".xlsx":
        await asyncio.to_thread(shutil.copyfile, path, disk_path)  # stored as-is, never parsed
    else:
        out_bytes = await asyncio.to_thread(_to_xlsx_bytes, ext, path)
        disk_path.write_bytes(out_bytes)
    return await repo.insert_file(user_id, session_id, out_name, str(disk_path), disk_path.stat().st_size)


async def delete_file(user_id: str, file_id: str) -> bool:
    """Delete a file's DB row AND its backing workbook on disk, so nothing is orphaned. The disk
    file is removed only when no OTHER file row still points to it."""
    logger.info("→ delete_file(user_id=%r file_id=%r)", user_id, file_id)  # autolog
    row = await repo.get_file_full(file_id, user_id)
    if not row:
        return False
    await repo.delete_file(file_id, user_id)

    dpath = row.get("disk_path")
    if dpath:
        remaining = await repo.list_for_session(row["session_id"])
        if not any(r.get("disk_path") == dpath for r in remaining):
            try:
                Path(dpath).unlink(missing_ok=True)
            except OSError as e:
                logger.warning("could not unlink %s: %s", dpath, e)
    return True


